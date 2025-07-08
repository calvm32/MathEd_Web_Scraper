import requests
import bs4
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Search Results"

column_index = 1 

while True:
    # Prompt user to enter a search
    text = input("Enter a search keyword or phrase (or type 'exit' to finish): ")
    
    if text.lower() == 'exit':
        break

    url = 'https://google.com/search?q=' + text
    request_result = requests.get(url)
    soup = bs4.BeautifulSoup(request_result.text, "html.parser")
    heading_objects = soup.find_all('h3')

    ws.cell(row=1, column=column_index, value=text)

    # Iterate through searches and write them to the sheet
    row_index = 2
    for heading in heading_objects:
        heading_text = heading.getText()
        link = heading.find_parent('a')
        if link:
            raw_url = link['href']
            url = raw_url.replace('/url?q=', '').split('&')[0]  # Clean up URL
        else:
            url = 'N/A'
        
        # Write the heading as a hyperlink
        cell = ws.cell(row=row_index, column=column_index, value=heading_text)
        cell.hyperlink = url
        cell.style = "Hyperlink"
        
        row_index += 1

    column_index += 1

# Save the workbook to a file
wb.save('search_results.xlsx')
