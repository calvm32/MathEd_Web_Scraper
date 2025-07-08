import asyncio
import aiohttp
import requests
import bs4
from openpyxl import Workbook
from urllib.parse import urlparse, parse_qs

async def fetch(session, url):
    try:
        async with session.get(url) as response:
            response.raise_for_status()
            return await response.text()
    except Exception as e:
        return f"Error: {e}"

async def process_heading(heading, session):
    heading_text = heading.getText()
    link = heading.find_parent('a')
    if link:
        raw_url = link['href']
        parsed_url = urlparse(raw_url)
        url = parse_qs(parsed_url.query).get('q')
        url = url[0] if url else 'N/A'
    else:
        url = 'N/A'

    domain_mentioned = "No"
    context = ""
    if url != 'N/A':
        page_text = await fetch(session, url)
        if not page_text.startswith("Error:"):
            page_soup = bs4.BeautifulSoup(page_text, "html.parser")
            page_text = page_soup.get_text()

            if 'domain' in page_text.lower():
                domain_mentioned = "Yes"
                start_index = page_text.lower().index('domain')
                end_index = min(len(page_text), start_index + 100)
                context = page_text[start_index:end_index]
        else:
            domain_mentioned = page_text
            context = ""
    
    return heading_text, url, domain_mentioned, context

async def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    column_index = 1

    while True:
        text = input("Enter a search keyword or phrase (or type 'exit' to finish): ")

        if text.lower() == 'exit':
            break

        url = 'https://google.com/search?q=' + text
        request_result = requests.get(url)
        soup = bs4.BeautifulSoup(request_result.text, "html.parser")
        heading_objects = soup.find_all('h3')[:5]

        ws.cell(row=1, column=column_index, value=text)
        ws.cell(row=2, column=column_index, value="Heading")
        ws.cell(row=2, column=column_index+1, value="URL")
        ws.cell(row=2, column=column_index+2, value="Domain Mentioned")
        ws.cell(row=2, column=column_index+3, value="Context")

        async with aiohttp.ClientSession() as session:
            tasks = [process_heading(heading, session) for heading in heading_objects]
            results = await asyncio.gather(*tasks)

        row_index = 3
        for heading_text, url, domain_mentioned, context in results:
            cell = ws.cell(row=row_index, column=column_index, value=heading_text)
            cell.hyperlink = url
            cell.style = "Hyperlink"
            ws.cell(row=row_index, column=column_index+1, value=url)
            ws.cell(row=row_index, column=column_index+2, value=domain_mentioned)
            ws.cell(row=row_index, column=column_index+3, value=context)
            row_index += 1

        column_index += 4

    wb.save('search_results.xlsx')

asyncio.run(main())
